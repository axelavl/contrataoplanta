"""
EmpleoEstado.cl — Verificador Masivo de URLs
Testea las 642 instituciones del repositorio y genera reporte Excel + JSON.

Uso:
    python verificar_urls.py                    # verifica todas
    python verificar_urls.py --sector Municipal # solo un sector
    python verificar_urls.py --limite 50        # primeras N instituciones
    python verificar_urls.py --workers 20       # concurrencia
"""

import json
import time
import asyncio
import argparse
import re
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

import aiohttp
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuración ──────────────────────────────────────────────────────
TIMEOUT_SEG     = 15       # segundos por request
MAX_WORKERS     = 30       # requests simultáneos (no abusar)
DELAY_ENTRE_REQ = 0.05     # delay mínimo entre requests
REINTENTOS      = 2        # reintentos en caso de error transitorio

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/119.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/118.0.0.0 Safari/537.36",
]

# ── Colores ────────────────────────────────────────────────────────────
C = {
    "azul":    "0B3D91", "azul_m": "1557C0", "azul_c": "D6E4F0",
    "verde":   "1B5E20", "verde_c":"E8F5E9",
    "rojo":    "9B2335", "rojo_c": "FDECEA",
    "naranja": "E65100", "naran_c":"FFF3E0",
    "amar_c":  "FFFDE7", "gris1":  "F4F6FB",
    "gris3":   "8A92A6", "blanco": "FFFFFF",
}

# ── Cargar repositorio ─────────────────────────────────────────────────
def cargar_repositorio(path: str) -> list[dict]:
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    return data["instituciones"]

# ── Verificación asíncrona ─────────────────────────────────────────────
async def verificar_url(session: aiohttp.ClientSession, url: str,
                        semaforo: asyncio.Semaphore, idx: int = 0) -> dict:
    """
    Verifica una URL y devuelve dict con resultado.
    """
    import random
    resultado = {
        "url": url,
        "status_code": None,
        "estado": None,          # OK / REDIRECCION / ERROR / TIMEOUT / INVALIDA
        "url_final": None,       # URL después de redirecciones
        "redireccion": False,
        "tiempo_ms": None,
        "error": None,
        "es_pagina_empleo": None, # si la página tiene contenido de empleo
    }

    # Validar formato de URL
    if not url or not url.startswith("http"):
        resultado["estado"] = "INVALIDA"
        resultado["error"] = "URL no comienza con http"
        return resultado

    # Limpiar URL
    url = url.strip()

    headers = {
        "User-Agent": USER_AGENTS[idx % len(USER_AGENTS)],
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "es-CL,es;q=0.9",
    }

    async with semaforo:
        await asyncio.sleep(DELAY_ENTRE_REQ * random.uniform(0.5, 1.5))

        for intento in range(REINTENTOS + 1):
            try:
                inicio = time.monotonic()
                async with session.get(
                    url, headers=headers,
                    timeout=aiohttp.ClientTimeout(total=TIMEOUT_SEG),
                    allow_redirects=True,
                    ssl=False  # ignorar errores SSL para mayor cobertura
                ) as resp:
                    duracion = int((time.monotonic() - inicio) * 1000)
                    resultado["status_code"] = resp.status
                    resultado["tiempo_ms"]   = duracion
                    resultado["url_final"]   = str(resp.url)
                    resultado["redireccion"] = str(resp.url) != url

                    if resp.status == 200:
                        resultado["estado"] = "OK"
                        # Leer primeros 5000 chars para detectar contenido de empleo
                        try:
                            texto = await resp.text(encoding="utf-8", errors="ignore")
                            texto_lower = texto[:5000].lower()
                            kw_empleo = ["concurso", "empleo", "postul", "trabaje",
                                         "cargo", "vacante", "convocatoria", "bases"]
                            resultado["es_pagina_empleo"] = any(k in texto_lower for k in kw_empleo)
                        except Exception:
                            resultado["es_pagina_empleo"] = None

                    elif resp.status in (301, 302, 303, 307, 308):
                        resultado["estado"] = "REDIRECCION"
                    elif resp.status == 403:
                        resultado["estado"] = "ERROR_403"
                    elif resp.status == 404:
                        resultado["estado"] = "ERROR_404"
                    elif resp.status == 500:
                        resultado["estado"] = "ERROR_500"
                    elif resp.status == 503:
                        resultado["estado"] = "ERROR_503"
                    else:
                        resultado["estado"] = f"HTTP_{resp.status}"

                    break  # éxito — salir del loop de reintentos

            except asyncio.TimeoutError:
                resultado["estado"] = "TIMEOUT"
                resultado["error"]  = f"Timeout ({TIMEOUT_SEG}s)"
                if intento < REINTENTOS:
                    await asyncio.sleep(2 ** intento)

            except aiohttp.ClientConnectorError as e:
                resultado["estado"] = "SIN_CONEXION"
                resultado["error"]  = str(e)[:120]
                if intento < REINTENTOS:
                    await asyncio.sleep(1)

            except aiohttp.ClientError as e:
                resultado["estado"] = "ERROR_CLIENTE"
                resultado["error"]  = str(e)[:120]
                break

            except Exception as e:
                resultado["estado"] = "ERROR"
                resultado["error"]  = str(e)[:120]
                break

    return resultado

async def verificar_todas(instituciones: list[dict],
                          workers: int = MAX_WORKERS) -> list[dict]:
    """Verifica sitio web y URL de empleo de cada institución."""
    semaforo = asyncio.Semaphore(workers)
    connector = aiohttp.TCPConnector(
        limit=workers,
        limit_per_host=3,  # máx 3 conexiones simultáneas por host
        ttl_dns_cache=300,
        force_close=False,
    )

    resultados = []
    total = len(instituciones)

    async with aiohttp.ClientSession(connector=connector) as session:
        # Crear tareas para sitio web y URL de empleo
        tareas = []
        meta = []

        for idx, inst in enumerate(instituciones):
            # Verificar sitio web
            tareas.append(verificar_url(session, inst["sitio_web"], semaforo, idx))
            meta.append({"id": inst["id"], "campo": "sitio_web", "inst": inst})

            # Verificar URL de empleo (si es diferente al sitio web)
            url_emp = inst["url_empleo"]
            if (url_emp and url_emp != inst["sitio_web"] and
                    url_emp not in ("empleospublicos.cl",
                                    "https://www.empleospublicos.cl")):
                tareas.append(verificar_url(session, url_emp, semaforo, idx))
                meta.append({"id": inst["id"], "campo": "url_empleo", "inst": inst})

        total_tareas = len(tareas)
        print(f"\n  Verificando {total_tareas} URLs ({total} instituciones)...")
        print(f"  Concurrencia: {workers} workers | Timeout: {TIMEOUT_SEG}s\n")

        # Ejecutar en lotes con progreso
        LOTE = 50
        completadas = 0

        for i in range(0, total_tareas, LOTE):
            lote_tareas = tareas[i:i+LOTE]
            lote_meta   = meta[i:i+LOTE]

            res_lote = await asyncio.gather(*lote_tareas, return_exceptions=True)

            for m, r in zip(lote_meta, res_lote):
                if isinstance(r, Exception):
                    r = {"url": m["inst"][m["campo"]], "estado": "ERROR",
                         "error": str(r)[:120], "status_code": None,
                         "tiempo_ms": None, "url_final": None,
                         "redireccion": False, "es_pagina_empleo": None}
                resultados.append({
                    "id":           m["id"],
                    "campo":        m["campo"],
                    "nombre":       m["inst"]["nombre"],
                    "sector":       m["inst"]["sector"],
                    "region":       m["inst"]["region"],
                    "url_original": m["inst"][m["campo"]],
                    **r
                })

            completadas += len(lote_tareas)
            pct = completadas / total_tareas * 100
            ok  = sum(1 for r in resultados if isinstance(r, dict) and r.get("estado") == "OK")
            print(f"  [{pct:5.1f}%] {completadas:>4}/{total_tareas} — OK: {ok}", end="\r")

    print(f"\n  Verificación completada.")
    return resultados

# ── Generar Excel de resultados ────────────────────────────────────────
def generar_excel(resultados: list[dict], instituciones: list[dict]) -> str:
    """Genera el Excel con resultados de verificación."""

    wb = openpyxl.Workbook()

    # Helpers
    def cel(ws, r, c, v, bg=C["blanco"], fg="1A2233", bold=False,
            sz=9, wrap=False, center=False):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(name="Calibri", bold=bold, color=fg, size=sz)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(
            horizontal="center" if center else "left",
            vertical="center", wrap_text=wrap)
        return cell

    def borde(ws, r, c):
        s = Side(style="thin", color="D4DAE8")
        ws.cell(r,c).border = Border(left=s, right=s, top=s, bottom=s)

    def color_estado(estado: str) -> tuple[str, str]:
        """Devuelve (bg, fg) según estado."""
        if estado == "OK":
            return C["verde_c"], C["verde"]
        elif estado in ("REDIRECCION", "HTTP_200"):
            return C["amar_c"], "856404"
        elif estado in ("TIMEOUT", "SIN_CONEXION", "ERROR_CLIENTE"):
            return C["naran_c"], C["naranja"]
        elif estado and estado.startswith("ERROR"):
            return C["rojo_c"], C["rojo"]
        elif estado == "INVALIDA":
            return "F0F0F0", C["gris3"]
        else:
            return C["rojo_c"], C["rojo"]

    # ── Organizar por institución ──
    por_id = {}
    for r in resultados:
        iid = r["id"]
        if iid not in por_id:
            por_id[iid] = {}
        por_id[iid][r["campo"]] = r

    # ── Hoja 1: Resultados completos ──
    ws1 = wb.active
    ws1.title = "Verificación Completa"
    ws1.sheet_view.showGridLines = False

    # Estadísticas rápidas
    total    = len(resultados)
    ok_count = sum(1 for r in resultados if r.get("estado") == "OK")
    err_count= sum(1 for r in resultados if r.get("estado") not in
                   ("OK", "REDIRECCION", "INVALIDA") and r.get("estado"))
    red_count= sum(1 for r in resultados if r.get("estado") == "REDIRECCION")
    emp_count= sum(1 for r in resultados
                   if r.get("campo") == "url_empleo" and r.get("es_pagina_empleo"))

    # Título
    NCOLS = 12
    ws1.merge_cells(f"A1:{get_column_letter(NCOLS)}1")
    c = ws1["A1"]
    c.value = f"VERIFICACIÓN DE URLs — Repositorio Instituciones Públicas Chile — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c.font = Font(name="Calibri", bold=True, size=13, color=C["blanco"])
    c.fill = PatternFill("solid", fgColor=C["azul"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 28

    # Barra de estadísticas
    stats = [
        (f"✓ OK: {ok_count}", C["verde_c"], C["verde"]),
        (f"↩ Redireccionadas: {red_count}", C["amar_c"], "856404"),
        (f"✗ Con error: {err_count}", C["rojo_c"], C["rojo"]),
        (f"Total URLs: {total}", C["azul_c"], C["azul"]),
        (f"Con contenido empleo: {emp_count}", C["verde_c"], C["verde"]),
    ]
    ws1.row_dimensions[2].height = 20
    ancho_col = NCOLS // len(stats)
    for i, (txt, bg, fg) in enumerate(stats):
        col_ini = 1 + i * ancho_col
        col_fin = col_ini + ancho_col - 1
        ws1.merge_cells(f"{get_column_letter(col_ini)}2:{get_column_letter(min(col_fin, NCOLS))}2")
        c = ws1.cell(2, col_ini, txt)
        c.font = Font(name="Calibri", bold=True, size=10, color=fg)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Cabeceras
    HDRS = [
        ("ID", 5), ("Institución", 38), ("Sector", 18), ("Región", 18),
        ("Campo verificado", 14), ("URL Original", 45), ("Estado", 16),
        ("HTTP", 6), ("URL Final", 40), ("Tiempo (ms)", 10),
        ("Contenido empleo", 14), ("Error / Detalle", 40),
    ]
    ws1.row_dimensions[3].height = 32
    for i, (h, w) in enumerate(HDRS, 1):
        cel(ws1, 3, i, h, bg=C["azul_m"], fg=C["blanco"],
            bold=True, sz=10, wrap=True, center=True)
        ws1.column_dimensions[get_column_letter(i)].width = w

    # Datos
    for idx, r in enumerate(sorted(resultados, key=lambda x: (x["id"], x["campo"]))):
        fila = 4 + idx
        bg_est, fg_est = color_estado(r.get("estado", ""))
        bg_fila = C["gris1"] if idx % 2 == 0 else C["blanco"]

        # Determinar si es sitio_web o url_empleo
        campo_lbl = "Sitio web" if r["campo"] == "sitio_web" else "Portal empleo"
        contenido = "✓ Sí" if r.get("es_pagina_empleo") else ("✗ No" if r.get("es_pagina_empleo") is False else "—")
        bg_cont   = C["verde_c"] if r.get("es_pagina_empleo") else (C["rojo_c"] if r.get("es_pagina_empleo") is False else C["blanco"])

        datos = [
            r["id"], r["nombre"], r["sector"], r["region"],
            campo_lbl, r.get("url_original",""),
            r.get("estado",""), r.get("status_code",""),
            r.get("url_final","") or "",
            r.get("tiempo_ms",""), contenido,
            r.get("error","") or "",
        ]
        bgs   = [bg_fila, bg_fila, bg_fila, bg_fila,
                 C["azul_c"] if r["campo"]=="url_empleo" else bg_fila,
                 bg_fila, bg_est, bg_est, bg_fila,
                 bg_fila, bg_cont, bg_fila]
        fgs   = ["1A2233"]*12
        fgs[6] = fg_est
        fgs[7] = fg_est
        cents  = [True,False,True,False,True,False,True,True,False,True,True,False]

        for j, (val, bgc, fgc, cent) in enumerate(zip(datos, bgs, fgs, cents), 1):
            cel(ws1, fila, j, val, bg=bgc, fg=fgc, center=cent,
                bold=(j==7), wrap=(j in [2,6,9,12]))
            borde(ws1, fila, j)

        ws1.row_dimensions[fila].height = 20

    ws1.freeze_panes = "A4"

    # ── Hoja 2: Solo errores (para acción) ──
    ws2 = wb.create_sheet("URLs con Error")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells(f"A1:{get_column_letter(NCOLS)}1")
    c = ws2["A1"]
    c.value = "URLs CON ERROR — Requieren corrección manual"
    c.font = Font(name="Calibri", bold=True, size=13, color=C["blanco"])
    c.fill = PatternFill("solid", fgColor=C["rojo"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    # Mismas cabeceras
    ws2.row_dimensions[2].height = 32
    for i, (h, w) in enumerate(HDRS, 1):
        cel(ws2, 2, i, h, bg=C["rojo"], fg=C["blanco"],
            bold=True, sz=10, wrap=True, center=True)
        ws2.column_dimensions[get_column_letter(i)].width = w

    errores = [r for r in resultados if r.get("estado") not in ("OK", "REDIRECCION", "INVALIDA")]
    for idx, r in enumerate(sorted(errores, key=lambda x: x.get("estado",""))):
        fila = 3 + idx
        bg_est, fg_est = color_estado(r.get("estado",""))
        bg_fila = C["gris1"] if idx % 2 == 0 else C["blanco"]
        campo_lbl = "Sitio web" if r["campo"] == "sitio_web" else "Portal empleo"
        contenido = "—"

        datos = [
            r["id"], r["nombre"], r["sector"], r["region"],
            campo_lbl, r.get("url_original",""),
            r.get("estado",""), r.get("status_code",""),
            r.get("url_final","") or "",
            r.get("tiempo_ms",""), contenido,
            r.get("error","") or "",
        ]
        for j, val in enumerate(datos, 1):
            cel(ws2, fila, j, val,
                bg=bg_est if j in [7,8] else bg_fila,
                fg=fg_est if j in [7,8] else "1A2233",
                center=(j in [1,5,7,8,10,11]),
                wrap=(j in [2,6,9,12]))
            borde(ws2, fila, j)

        ws2.row_dimensions[fila].height = 20

    if not errores:
        ws2.cell(3, 1, "✓ No se encontraron URLs con error").font = \
            Font(name="Calibri", bold=True, color=C["verde"], size=12)

    ws2.freeze_panes = "A3"

    # ── Hoja 3: Sin contenido de empleo (URL empleo incorrecta) ──
    ws3 = wb.create_sheet("Sin Contenido Empleo")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells(f"A1:{get_column_letter(NCOLS)}1")
    c = ws3["A1"]
    c.value = "URLs PORTAL EMPLEO — Sin contenido de empleo detectado (probablemente URL incorrecta)"
    c.font = Font(name="Calibri", bold=True, size=12, color=C["blanco"])
    c.fill = PatternFill("solid", fgColor=C["naranja"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28

    ws3.row_dimensions[2].height = 32
    for i, (h, w) in enumerate(HDRS, 1):
        cel(ws3, 2, i, h, bg=C["naranja"], fg=C["blanco"],
            bold=True, sz=10, wrap=True, center=True)
        ws3.column_dimensions[get_column_letter(i)].width = w

    sin_empleo = [r for r in resultados
                  if r["campo"] == "url_empleo"
                  and r.get("estado") == "OK"
                  and r.get("es_pagina_empleo") is False]

    for idx, r in enumerate(sin_empleo):
        fila = 3 + idx
        bg_fila = C["gris1"] if idx % 2 == 0 else C["blanco"]
        datos = [
            r["id"], r["nombre"], r["sector"], r["region"],
            "Portal empleo", r.get("url_original",""),
            r.get("estado",""), r.get("status_code",""),
            r.get("url_final","") or "",
            r.get("tiempo_ms",""), "✗ No",
            "URL responde 200 pero sin contenido de empleo detectado",
        ]
        for j, val in enumerate(datos, 1):
            cel(ws3, fila, j, val, bg=C["naran_c"] if j == 11 else bg_fila,
                center=(j in [1,5,7,8,10,11]),
                wrap=(j in [2,6,9,12]))
            borde(ws3, fila, j)
        ws3.row_dimensions[fila].height = 20

    ws3.freeze_panes = "A3"

    # ── Hoja 4: Resumen estadístico ──
    ws4 = wb.create_sheet("Resumen")
    ws4.sheet_view.showGridLines = False

    ws4["A1"].value = "RESUMEN DE VERIFICACIÓN"
    ws4["A1"].font = Font(name="Calibri", bold=True, size=14, color=C["blanco"])
    ws4["A1"].fill = PatternFill("solid", fgColor=C["azul"])
    ws4.merge_cells("A1:D1")
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 28

    sectores_err = {}
    for r in resultados:
        s = r["sector"]
        if s not in sectores_err:
            sectores_err[s] = {"total":0,"ok":0,"error":0,"sin_contenido":0}
        sectores_err[s]["total"] += 1
        if r.get("estado") == "OK":
            sectores_err[s]["ok"] += 1
            if r["campo"] == "url_empleo" and r.get("es_pagina_empleo") is False:
                sectores_err[s]["sin_contenido"] += 1
        elif r.get("estado") not in ("REDIRECCION","INVALIDA"):
            sectores_err[s]["error"] += 1

    for i, (h, w) in enumerate([("Sector",28),("URLs OK",12),("Con error",12),("Sin contenido empleo",22)], 1):
        cel(ws4, 2, i, h, bg=C["azul_m"], fg=C["blanco"], bold=True, center=True)
        ws4.column_dimensions[get_column_letter(i)].width = w

    for idx, (sect, st) in enumerate(sorted(sectores_err.items())):
        fila = 3 + idx
        bg = C["gris1"] if idx%2==0 else C["blanco"]
        bg_err = C["rojo_c"] if st["error"] > 0 else C["verde_c"]
        bg_sinc= C["naran_c"] if st["sin_contenido"] > 0 else C["verde_c"]
        for j, val in enumerate([sect, st["ok"], st["error"], st["sin_contenido"]], 1):
            cel(ws4, fila, j, val,
                bg=bg if j==1 else (bg_err if j==3 else (bg_sinc if j==4 else bg)),
                bold=(j==1), center=(j>1))
            borde(ws4, fila, j)

    # Guardar
    ruta = "/mnt/user-data/outputs/verificacion_urls_completa.xlsx"
    wb.save(ruta)
    return ruta

# ── Generar JSON de resultados ─────────────────────────────────────────
def generar_json_resultados(resultados: list[dict],
                             instituciones: list[dict]) -> str:
    """Genera JSON con resultados y repositorio actualizado."""

    # Crear dict de resultados por id+campo
    res_dict = {}
    for r in resultados:
        key = f"{r['id']}_{r['campo']}"
        res_dict[key] = r

    # Actualizar estado de verificación en instituciones
    insts_actualizadas = []
    for inst in instituciones:
        inst_copia = dict(inst)

        res_web = res_dict.get(f"{inst['id']}_sitio_web")
        res_emp = res_dict.get(f"{inst['id']}_url_empleo")

        if res_web:
            inst_copia["verificacion_sitio_web"] = {
                "estado": res_web.get("estado"),
                "status_code": res_web.get("status_code"),
                "tiempo_ms": res_web.get("tiempo_ms"),
                "url_final": res_web.get("url_final"),
            }

        if res_emp:
            inst_copia["verificacion_url_empleo"] = {
                "estado": res_emp.get("estado"),
                "status_code": res_emp.get("status_code"),
                "tiempo_ms": res_emp.get("tiempo_ms"),
                "url_final": res_emp.get("url_final"),
                "tiene_contenido_empleo": res_emp.get("es_pagina_empleo"),
            }

        # Calcular estado general
        ok_web = res_web and res_web.get("estado") == "OK"
        ok_emp = res_emp and res_emp.get("estado") == "OK"

        if ok_web and ok_emp:
            inst_copia["estado_verificacion"] = "Verificada OK"
        elif ok_web and not res_emp:
            inst_copia["estado_verificacion"] = "Sitio OK — URL empleo = empleospublicos.cl"
        elif ok_web:
            inst_copia["estado_verificacion"] = "Sitio OK — URL empleo con error"
        elif not res_web:
            inst_copia["estado_verificacion"] = "No verificado"
        else:
            inst_copia["estado_verificacion"] = "Error en sitio web"

        insts_actualizadas.append(inst_copia)

    # Estadísticas
    ok_total  = sum(1 for r in resultados if r.get("estado") == "OK")
    err_total = sum(1 for r in resultados if r.get("estado") not in
                    ("OK", "REDIRECCION", "INVALIDA") and r.get("estado"))

    output = {
        "metadata": {
            "nombre": "Repositorio Verificado de Instituciones Públicas de Chile",
            "proyecto": "EmpleoEstado.cl",
            "version": "1.1",
            "fecha_verificacion": datetime.now().isoformat(),
            "total_instituciones": len(instituciones),
            "total_urls_verificadas": len(resultados),
            "estadisticas": {
                "urls_ok":          ok_total,
                "urls_con_error":   err_total,
                "urls_redireccion": sum(1 for r in resultados if r.get("estado")=="REDIRECCION"),
                "tasa_exito_pct":   round(ok_total/len(resultados)*100, 1) if resultados else 0,
            }
        },
        "instituciones": insts_actualizadas
    }

    ruta = "/mnt/user-data/outputs/repositorio_verificado.json"
    with open(ruta, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    return ruta

# ── Main ───────────────────────────────────────────────────────────────
async def main():
    parser = argparse.ArgumentParser(description="Verificador masivo de URLs instituciones públicas Chile")
    parser.add_argument("--repo",    default="/mnt/user-data/outputs/repositorio_instituciones_publicas_chile.json",
                        help="Ruta al JSON del repositorio")
    parser.add_argument("--sector",  default=None, help="Filtrar por sector")
    parser.add_argument("--limite",  type=int, default=None, help="Limitar N instituciones")
    parser.add_argument("--workers", type=int, default=MAX_WORKERS)
    args = parser.parse_args()

    print("\n" + "="*60)
    print("  EmpleoEstado.cl — Verificador Masivo de URLs")
    print("="*60)

    # Cargar
    instituciones = cargar_repositorio(args.repo)
    print(f"\n  Repositorio cargado: {len(instituciones)} instituciones")

    # Filtrar
    if args.sector:
        instituciones = [i for i in instituciones if args.sector.lower() in i["sector"].lower()]
        print(f"  Filtrado por sector '{args.sector}': {len(instituciones)} instituciones")

    if args.limite:
        instituciones = instituciones[:args.limite]
        print(f"  Limitado a primeras {args.limite} instituciones")

    inicio = time.time()

    # Verificar
    resultados = await verificar_todas(instituciones, workers=args.workers)

    duracion = time.time() - inicio
    print(f"\n  Duración total: {duracion:.1f}s")

    # Estadísticas
    ok    = sum(1 for r in resultados if r.get("estado") == "OK")
    err   = sum(1 for r in resultados if r.get("estado") not in ("OK","REDIRECCION","INVALIDA"))
    redir = sum(1 for r in resultados if r.get("estado") == "REDIRECCION")
    total = len(resultados)

    print(f"\n  ─── RESULTADOS ───────────────────────────────")
    print(f"  Total URLs verificadas: {total}")
    print(f"  ✓ OK:                   {ok}  ({ok/total*100:.1f}%)")
    print(f"  ↩ Redireccionadas:      {redir}  ({redir/total*100:.1f}%)")
    print(f"  ✗ Con error:            {err}  ({err/total*100:.1f}%)")
    print(f"  ─────────────────────────────────────────────")

    # Generar archivos
    print("\n  Generando Excel de resultados...")
    ruta_xlsx = generar_excel(resultados, instituciones)
    print(f"  Excel: {ruta_xlsx}")

    print("  Generando JSON verificado...")
    ruta_json = generar_json_resultados(resultados, instituciones)
    print(f"  JSON:  {ruta_json}")

    print("\n" + "="*60)
    print("  ¡Verificación completada!")
    print("="*60 + "\n")

if __name__ == "__main__":
    asyncio.run(main())
